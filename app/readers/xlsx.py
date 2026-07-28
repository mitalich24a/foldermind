"""
FolderMind
Copyright (c) 2026 Mitali Choubisa.
All rights reserved.

Reader for Excel documents.
"""

from pathlib import Path

from openpyxl import load_workbook

from app.models.document import Document
from app.readers.base import DocumentReader


class XLSXReader(DocumentReader):
    """
    Reader for Excel (.xlsx) files.
    """

    def read(self, document: Document) -> str:
        path = Path(document.path)

        workbook = load_workbook(
            filename=path,
            data_only=True,
        )

        lines = []

        for sheet in workbook.worksheets:
            lines.append(f"Sheet: {sheet.title}")

            for row in sheet.iter_rows(values_only=True):
                values = [
                    str(cell)
                    for cell in row
                    if cell is not None
                ]

                if values:
                    lines.append(", ".join(values))

            lines.append("")

        return "\n".join(lines)